"""Excel tracker for job applications."""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

COLUMNS = [
    "Job ID", "Date Found", "Date Applied", "Platform", "Company",
    "Role", "Location", "Remote", "Salary", "Score (%)",
    "Matching Skills", "Missing Skills", "Status",
    "Application URL", "Contact Person", "Contact Email",
    "LinkedIn Network Match", "Notes"
]

STATUS_COLORS = {
    "applied": "C6EFCE",
    "manual_required": "FFEB9C",
    "skipped": "F2F2F2",
    "rejected": "FFC7CE",
    "interview": "BDD7EE",
    "new": "FFFFFF",
}


class ExcelTracker:
    """Excel tracker for job applications."""

    def __init__(self, filepath: str = "data/jobs_tracker.xlsx"):
        self.filepath = filepath
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

    def update_job(self, job: dict) -> bool:
        """Add or update job in Excel tracker."""
        try:
            # Load or create workbook
            if os.path.exists(self.filepath):
                wb = load_workbook(self.filepath)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Job Applications"
                self._write_header(ws)

            # Check if job exists by ID
            existing_row = None
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == job.get("id"):
                    existing_row = row
                    break

            if existing_row:
                row_num = existing_row
            else:
                row_num = ws.max_row + 1

            # Prepare row data
            row_data = [
                job.get("id"),
                job.get("date_found", ""),
                job.get("date_applied", ""),
                job.get("platform", ""),
                job.get("company", ""),
                job.get("title", ""),
                job.get("location", ""),
                "Yes" if job.get("is_remote") else "No",
                job.get("salary", "N/A"),
                job.get("score", ""),
                ", ".join(job.get("matching_skills", [])),
                ", ".join(job.get("missing_skills", [])),
                job.get("status", "new"),
                job.get("application_url", ""),
                job.get("contact_person", ""),
                job.get("contact_email", ""),
                job.get("linkedin_network_match", ""),
                job.get("notes", ""),
            ]

            # Write row
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                # Color status column
                if col == 13:
                    status = job.get("status", "new")
                    color = STATUS_COLORS.get(status, "FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=color)

            wb.save(self.filepath)
            return True
        except Exception as e:
            print(f"Error updating Excel tracker: {e}")
            return False

    def _write_header(self, ws):
        """Write header row."""
        for col, header in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def get_all_jobs(self) -> list[dict]:
        """Read all jobs from Excel."""
        if not os.path.exists(self.filepath):
            return []

        try:
            wb = load_workbook(self.filepath)
            ws = wb.active
            jobs = []

            for row in range(2, ws.max_row + 1):
                job = {}
                for col, header in enumerate(COLUMNS, 1):
                    value = ws.cell(row=row, column=col).value
                    job[header] = value

                jobs.append(job)

            return jobs
        except Exception as e:
            print(f"Error reading Excel tracker: {e}")
            return []
